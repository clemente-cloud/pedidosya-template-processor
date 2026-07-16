from datetime import date
from io import BytesIO
from pathlib import Path

import openpyxl
from fastapi import FastAPI, File, HTTPException, UploadFile
from fastapi.responses import Response
from fastapi.staticfiles import StaticFiles

from . import excel_writer, rules, session_store
from .models import DownloadRequest, Producto, Stats, UploadResponse

BASE_DIR = Path(__file__).resolve().parent.parent.parent
FRONTEND_DIR = BASE_DIR / "frontend"

PRIO_SHEETS = ["MENU", "TEMPLATE", "CATALOGO", "PRODUCTOS", "CATALOG"]

# Sin CORSMiddleware a proposito: el frontend siempre lo sirve este mismo
# proceso FastAPI (StaticFiles mas abajo), nunca otro origen, asi que no hay
# pedidos cross-origin legitimos que habilitar.
app = FastAPI(title="PedidosYa Template Processor")


def elegir_hoja(wb):
    for name in wb.sheetnames:
        if name.upper() in PRIO_SHEETS:
            return name
    return wb.sheetnames[0]


@app.get("/api/health")
async def health():
    return {"status": "ok"}


@app.post("/api/upload", response_model=UploadResponse)
async def upload(file: UploadFile = File(...)):
    if not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(400, "El archivo debe ser .xlsx o .xls")

    contents = await file.read()
    try:
        wb = openpyxl.load_workbook(BytesIO(contents), data_only=True, read_only=True)
    except Exception as exc:  # noqa: BLE001 - reportamos cualquier error de lectura al cliente
        raise HTTPException(400, f"No se pudo leer el archivo: {exc}") from exc

    sheet_name = elegir_hoja(wb)
    ws = wb[sheet_name]
    rows = [list(row) for row in ws.iter_rows(values_only=True)]
    wb.close()

    productos = rules.procesar_filas(rows)

    total = len(productos)
    ok = sum(1 for p in productos if p["estado"] == "ok")
    warn = sum(1 for p in productos if p["estado"] == "warn")
    err = sum(1 for p in productos if p["estado"] == "err")
    desc_larga = sum(1 for p in productos if any("250" in w for w in p["warns"]))
    # Codigo de barras vacio no es un error (es opcional), pero se avisa en
    # un popup aparte para que el KAM lo tenga presente. Imagen si es
    # obligatoria: se cuenta a partir del error real generado en rules.py.
    sin_barcode = sum(1 for p in productos if not p["ean"])
    sin_imagen = sum(1 for p in productos if any("imagen" in e.lower() for e in p["errs"]))

    session_id = session_store.create_session(file.filename, contents, sheet_name, productos)

    return UploadResponse(
        session_id=session_id,
        filename=file.filename,
        productos=[Producto(**p) for p in productos],
        stats=Stats(
            total=total, ok=ok, warn=warn, err=err, desc_larga=desc_larga,
            sin_barcode=sin_barcode, sin_imagen=sin_imagen,
        ),
    )


@app.post("/api/download/{session_id}")
async def download(session_id: str, body: DownloadRequest):
    session = session_store.get_session(session_id)
    if session is None:
        raise HTTPException(404, "Sesión no encontrada o expirada. Volvé a subir el archivo.")

    data = excel_writer.generar_descarga(
        session["original_bytes"],
        session["sheet_name"],
        [p.model_dump() for p in body.productos],
    )

    filename = f"template_procesado_{date.today().isoformat()}.xlsx"
    return Response(
        content=data,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# Debe ir despues de las rutas /api/*: sirve frontend/ como estaticos y hace de
# fallback para "/" y cualquier otra ruta no reconocida.
app.mount("/", StaticFiles(directory=FRONTEND_DIR, html=True), name="frontend")
