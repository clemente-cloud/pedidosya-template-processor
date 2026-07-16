"""Genera la descarga inyectando los datos editados en el template original.

A diferencia del enfoque anterior (SheetJS armando un libro nuevo en blanco),
esto abre con openpyxl el archivo tal cual lo subio el usuario y solo
sobreescribe los valores de las columnas de datos (A-J, L-O) desde la fila 3.
Nunca se toca la hoja "Instrucciones y Ejemplos", ni las columnas K y P
(formulas), ni estilos, colores, anchos de columna o validaciones — todo eso
queda intacto porque nunca se reconstruye el libro, solo se editan celdas
puntuales sobre el original.
"""
from io import BytesIO

import openpyxl

# Columna (1-based) de cada campo editable. K (11) y P (16) son formulas y
# nunca se escriben.
COLUMNAS = {
    "activo": 1, "ean": 2, "sku": 3, "precio": 4, "sec": 5,
    "que_es": 6, "marca": 7, "variante": 8, "cont": 9, "uni": 10,
    "img": 12, "desc": 13, "impuestos": 14,
}

COLUMNAS_EAN = {"ean": 2, "ean_fraccionado": 15}


def _valor(v):
    """Convierte '' / None en None (celda realmente vacia), para que las
    formulas de Excel que usan ISBLANK() se comporten igual que en un
    template llenado a mano."""
    if v is None:
        return None
    if isinstance(v, str) and v.strip() == "":
        return None
    return v


def _escribir_ean(ws, row, col, valor):
    v = _valor(valor)
    cell = ws.cell(row=row, column=col)
    if v is None:
        cell.value = None
        return
    s = str(v).strip()
    if s.isdigit():
        cell.value = int(s)
        cell.number_format = "0"
    else:
        cell.value = s


def generar_descarga(original_bytes: bytes, sheet_name: str, productos: list[dict]) -> bytes:
    wb = openpyxl.load_workbook(BytesIO(original_bytes))
    ws = wb[sheet_name]

    for i, p in enumerate(productos):
        row = i + 3
        for field, col in COLUMNAS.items():
            ws.cell(row=row, column=col).value = _valor(p.get(field))
        for field, col in COLUMNAS_EAN.items():
            _escribir_ean(ws, row, col, p.get(field))

    # openpyxl no evalua formulas: forzamos que Excel recalcule K y P al
    # abrir el archivo, ya que sus valores cacheados quedaron desactualizados.
    wb.calculation.fullCalcOnLoad = True

    out = BytesIO()
    wb.save(out)
    return out.getvalue()
